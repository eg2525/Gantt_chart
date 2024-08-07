import streamlit as st
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter
import io

# Streamlitアプリの開始
st.title('ガントチャート作成アプリ')

def load_data(file):
    df = pd.read_csv(file, encoding='cp932')
    columns_to_drop = [
        'レコードの開始行', '標題', '担当社員', '担当者', 'レコード番号', '更新者', '作成者', '更新日時', '作成日時',
        'ステータス', 'プロジェクトコード', '関連者', '削除依頼', '削除依頼者', '削除依頼理由', '削除依頼日',
        'ルックアップ(被相続人)', '被相続人:顧客コード', '顧客名&ﾌﾘｶﾞﾅ', 'サーバーアドレス', 'ルックアップ(相続人)',
        '相続人:顧客名', '相続人:顧客コード', '作業予定者', '総予定工数', '工程リスト', '解約事由', '解約日'
    ]
    df = df.drop(columns_to_drop, axis=1)
    df['開始予定日'] = pd.to_datetime(df['開始予定日'], errors='coerce')
    df['終了予定日'] = pd.to_datetime(df['終了予定日'], errors='coerce')
    df['相続開始日'] = pd.to_datetime(df['相続開始日'], errors='coerce')
    return df

def create_gantt_chart(df, selected_tasks):
    inheritance_start = df['相続開始日'].min()
    calendar_start = inheritance_start - pd.to_timedelta(inheritance_start.weekday(), unit='D')
    calendar_end = df['終了予定日'].max()
    calendar_days = pd.date_range(start=calendar_start, end=calendar_end, freq='D')

    wb = Workbook()
    ws = wb.active
    ws.title = 'ガントチャート'

    blue_color = '87CEFA'
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    month_groups = calendar_days.to_series().groupby(calendar_days.to_period("M"))

    start_row = 1

    for month, days in month_groups:
        ws.cell(row=start_row, column=1, value='作業名')
        for i, day in enumerate(days, start=2):
            cell = ws.cell(row=start_row, column=i, value=day.strftime('%Y-%m-%d'))
            apply_styles(cell, bold=True, border=thin_border, alignment=Alignment(horizontal='center', vertical='center'))
            ws.column_dimensions[get_column_letter(i)].width = 15

        filtered_df = df[df['工程'].isin(selected_tasks)]
        task_rows = {task: idx + start_row + 1 for idx, task in enumerate(filtered_df['作業名'].unique())}
        for task, row in task_rows.items():
            cell = ws.cell(row=row, column=1, value=task)
            apply_styles(cell, bold=True, border=thin_border, alignment=Alignment(horizontal='center', vertical='center'))
            ws.row_dimensions[row].height = 20

        filtered_df = filtered_df.dropna(subset=['開始予定日', '終了予定日'])
        for _, row in filtered_df.iterrows():
            task_row = task_rows[row['作業名']]
            start_col = (row['開始予定日'] - days[0]).days + 2
            end_col = (row['終了予定日'] - days[0]).days + 2

            if start_col < 2:
                start_col = 2
            if end_col > len(days) + 1:
                end_col = len(days) + 1

            apply_task_colors(ws, task_row, start_col, end_col, blue_color, thin_border)

        start_row += len(task_rows) + 2

    adjust_column_width(ws)
    return wb

def apply_styles(cell, bold=False, border=None, alignment=None):
    if bold:
        cell.font = Font(bold=True)
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment

def apply_task_colors(ws, task_row, start_col, end_col, color, border):
    for i in range(start_col, end_col + 1):
        cell = ws.cell(row=task_row, column=i)
        cell.border = border
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

def adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max((max_length + 2), 20) if column == 'A' else (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    ws.column_dimensions['A'].width = max(ws.column_dimensions['A'].width, 20)

def save_excel(wb):
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

uploaded_file = st.file_uploader("ファイルをアップロードしてください", type=["csv"])

if uploaded_file is not None:
    df = load_data(uploaded_file)
    unique_tasks = df['工程'].unique()
    selected_tasks = st.multiselect('表示する工程を選択してください', unique_tasks, default=list(unique_tasks))

    if st.button('ガントチャート作成'):
        wb = create_gantt_chart(df, selected_tasks)
        output = save_excel(wb)
        st.download_button(label='ガントチャートをダウンロード', data=output, file_name='GanttChart.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
